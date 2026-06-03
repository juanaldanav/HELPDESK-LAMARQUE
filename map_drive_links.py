# Replica EXACTA del recorrido de generate_csvs.py para mapear folio LMQ -> link Drive (Evidencia/Ficha).
# Salida: SQL (stdout) que llena lmq_tmp_links y actualiza comment de cada tabla de activos.
# Uso: PYTHONIOENCODING=utf-8 python3 soporte-v2/map_drive_links.py > /tmp/drive_links.sql
import openpyxl, sys
from collections import defaultdict

EXCEL = 'INVENTARIO DE ACTIVOS BORRADOR.xlsx'
SUC_TAG = {
    'OFICINA':'OFC','CEDIS':'CDS','CHAPULE REPOSTERIA':'CHR','CHAPULE DRIVE':'CHD',
    'QUINTAS':'QNT','VALLE ALTO':'VAL','GUADALUPE':'GDL','PEDRO INFANTE':'PDI',
    'PRIVANZAS':'PRV','EXPLANADA':'EXP','TRES RÍOS':'TRS','CUATRO RÍOS':'CRS',
    'PRIMAVERA':'PRM','CONQUISTA DRIVE':'CQD','CONQUISTA PLAZA':'CQP','CAMPUS DIGITAL':'CAM',
}
TYP_CODE = {
    'Refrigeración':'REF','Maquinaria de Café / Preparación':'MCF',
    'Equipo de Cómputo / POS':'CPT','Electrónica y Seguridad':'ESG',
    'Mobiliario':'MOB','Utensilios y Menores':'UTL',
    'Televisiones y Señalización':'TV','Jardinería y Exterior':'JRD',
}

def infer_tipo(eq):
    if not eq: return 'Utensilios y Menores'
    e = str(eq).upper()
    if any(k in e for k in ['REFRIGER','CONGELAD','CUARTO DE C','VITRINA','FRIGORI',
        'FREEZER','FABRICAD','HIELO','MINISPLI','ACONDICION','SISTEMA DE AIRE',
        'SPLIT','CLIMA']): return 'Refrigeración'
    if any(k in e for k in ['CAFETERA','HORNO','BATIDOR','AMASAD','MICROOND',
        'LICUAD','MAQUINA AL','MÁQUINA AL','MOTOR','PROCESAD','SELLAD','ESTUFA',
        'CUCHIMICK','MOLINO','BASCULA','BÁSCULA','TRAMPA','VAPOR','GALLETA',
        'FREIDORA','PLANCHA','PARRILLA','CAMPANA']): return 'Maquinaria de Café / Preparación'
    if any(k in e for k in ['MONITOR','CPU','LAPTOP','IPAD','TABLET','TECLADO',
        'MOUSE','MAUSE','IMPRESORA','CELULAR','TERMINAL','APEXA','SCANNER',
        'ROUTER','SWITCH','NO BREAK','UPS']): return 'Equipo de Cómputo / POS'
    if any(k in e for k in ['CÁMARA','CAMARA','BOCINA','EXTINTOR','ALARMA',
        'LECTOR','TELÉFONO','TELEFONO','RELOJ','RADIO','KIOSKO','GABINETE NEG',
        'CAJA FUERTE','INTERFON','DVR','NVR']): return 'Electrónica y Seguridad'
    if any(k in e for k in ['TELEVISIÓN','TELEVISION','PANTALLA','TV ','SMART TV',
        'DISPLAY']): return 'Televisiones y Señalización'
    if any(k in e for k in ['MESA','SILLA','ESCRITOR','MUEBLE','ARCHIVER',
        'GABINETE','ESTANTE','PIZARR','BARRA','PERIQUERA','BANCO','REPISA',
        'RACK','LOKER','ESPIGUERO','BOTE','CARRITO','SILLON','SOFA']): return 'Mobiliario'
    if any(k in e for k in ['MACETA','PLANTA','JARDIN','PASTO','ARBOL']): return 'Jardinería y Exterior'
    return 'Utensilios y Menores'

LOTE_KW = ['TAZA','CUCHARA','VASO','BANDEJA','MOLDE','RECIPIENTE','CAJA DE',
           'ROLLO','BOLSA','SERVILLETA','GUANTE','TOALLA','PLATO','TENEDOR',
           'CUCHILLO','VIDRIO','COPA','CHAROL']
def is_consumable(eq, qty):
    if qty > 10: return True
    e = str(eq).upper()
    return any(k in e for k in LOTE_KW)

SKIP = {'ANEXOS DEL CATÁLOGO', 'MANTENIMEINTO'}

counters = defaultdict(int)
def gen_folio(tag, tipo):
    code = TYP_CODE.get(tipo, 'UTL')
    key = f"{tag}-{code}"
    counters[key] += 1
    return f"LMQ-{tag}-{code}-{counters[key]:04d}"

# Cargar dos veces: con data_only para valores (igual que el original) y sin para hyperlinks
wb = openpyxl.load_workbook(EXCEL, data_only=True)
mapping = []  # (folio, url)

for sn in wb.sheetnames:
    sn_norm = sn.strip().upper()
    if sn_norm in SKIP or sn_norm.startswith('ANEXOS'): continue
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
    if not rows: continue
    hdr = [c.value for c in rows[0]]

    def col(keywords):
        for i, hh in enumerate(hdr):
            if hh and any(k.lower() in str(hh).lower() for k in keywords):
                return i
        return None

    ei = col(['Equipo'])
    qi = col(['Cantidad'])
    ni = col(['Evidencia'])
    hi = col(['Ficha'])
    if ei is None: continue

    tag = SUC_TAG.get(sn_norm, 'GEN')

    for row in rows[1:]:
        eq = row[ei].value if ei < len(row) else None
        if not eq or str(eq).strip() in ('', 'nan', 'None'): continue
        try:
            qv = row[qi].value if qi is not None and qi < len(row) else 1
            qty = max(1, int(float(str(qv or 1).replace('-', '0') or 1)))
        except Exception:
            qty = 1
        eq_str = str(eq).strip().title()
        if is_consumable(eq_str, qty):
            continue
        # link de la fila: Evidencia (N) preferido, luego Ficha (H)
        url = None
        for ci in (ni, hi):
            if ci is not None and ci < len(row) and row[ci].hyperlink is not None:
                url = str(row[ci].hyperlink.target); break
            if ci is not None and ci < len(row) and row[ci].value and 'http' in str(row[ci].value):
                url = str(row[ci].value).strip(); break
        n_units = (1 if qty > 10 else qty)
        for _ in range(n_units):
            folio = gen_folio(tag, infer_tipo(eq_str))
            if url:
                mapping.append((folio, url.split()[0]))

print(f"-- folios con link: {len(mapping)}", file=sys.stderr)
print("SET NAMES utf8mb4;")
print("DROP TABLE IF EXISTS lmq_tmp_links;")
print("CREATE TABLE lmq_tmp_links (serial VARCHAR(64) PRIMARY KEY, url VARCHAR(500)) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;")
for folio, url in mapping:
    u = url.replace("\\", "").replace("'", "''")[:490]
    print(f"INSERT IGNORE INTO lmq_tmp_links VALUES ('{folio}', '{u}');")
TABLES = [
    'glpi_plugin_genericobject_refrigeracions','glpi_plugin_genericobject_maquinariadecafes',
    'glpi_plugin_genericobject_mobiliarios','glpi_plugin_genericobject_electronicayseguridads',
    'glpi_plugin_genericobject_jardinerias','glpi_plugin_genericobject_televisiones',
    'glpi_plugin_genericobject_utensilios','glpi_plugin_genericobject_herramientasoportes',
    'glpi_computers',
]
for t in TABLES:
    print(f"UPDATE {t} a JOIN lmq_tmp_links l ON l.serial = a.serial "
          f"SET a.comment = CONCAT(COALESCE(a.comment,''), ' | Ref: ', l.url) "
          f"WHERE COALESCE(a.comment,'') NOT LIKE '%drive.google%';")
print("SELECT (SELECT COUNT(*) FROM lmq_tmp_links) AS folios_con_link;")
