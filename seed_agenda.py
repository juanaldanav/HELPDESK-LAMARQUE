# Genera SQL para sembrar lmq_agenda desde Cronograma_Mantenimientos_2026.xlsx
# Uso: python seed_agenda.py > seed_agenda.sql   (correr desde D:\LAMARQUE\GLPI)
import openpyxl, datetime, re, sys

WB = 'Cronograma_Mantenimientos_2026.xlsx'
ENTITY = {  # nombre en Excel -> entities_id glpidb
    'Campus Digital': 16, 'Guadalupe': 7, 'Tres Rios': 11, '4 Rios': 12,
    'Chapultepec Drive Thru': 4, 'Primavera': 13, 'Privanzas': 9,
    'Pedro Infante': 8, 'Conquista': 15, 'Conquista Drive Thru': 14,
    'Explanada': 10, 'Valle Alto': 6, 'Quintas': 5, 'Chapultepec': 3,
}
MES = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
MESTXT = {'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,
          'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12}

def parse_cell(v):
    """-> lista de (fecha, fecha_fin|None). Vacio si N/A o ilegible."""
    if v is None: return []
    if isinstance(v, (datetime.datetime, datetime.date)):
        return [(v.strftime('%Y-%m-%d'), None)]
    s = str(v).strip()
    if not s or s.upper() == 'N/A': return []
    m = re.findall(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        d = [f'{y}-{int(mo):02d}-{int(da):02d}' for da, mo, y in m]
        return [(d[0], d[1] if len(d) > 1 else None)]
    # "Miercoles 11 de febrero" / "Miercoles 11 y 27 de febrero"
    m2 = re.search(r'(\d{1,2})(?:\s*y\s*(\d{1,2}))?\s*de\s*([a-zA-Záéíóú]+)', s, re.I)
    if m2:
        mes = MESTXT.get(m2.group(3).lower())
        if mes:
            out = [(f'2026-{mes:02d}-{int(m2.group(1)):02d}', None)]
            if m2.group(2): out.append((f'2026-{mes:02d}-{int(m2.group(2)):02d}', None))
            return out
    print(f'-- AVISO: celda ilegible omitida: {s!r}', file=sys.stderr)
    return []

wb = openpyxl.load_workbook(WB, data_only=True)
ws = wb['Calendario Mtto 2026']
rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

events = []
i = 0
while i < len(rows):
    a = rows[i][0]
    title = str(a).strip() if a else ''
    is_block = title and (rows[i][1] is None or str(rows[i][1]).strip() == '') and ('2026' in title or 'Marino' in title)
    if is_block:
        tipo = title.replace(' - 2026', '').strip()
        if 'Marino' in tipo: tipo = 'Café Marino'
        hdr = rows[i + 1] if i + 1 < len(rows) else None
        colmes = {ci: MES[str(h).strip()] for ci, h in enumerate(hdr or []) if h and str(h).strip() in MES}
        j = i + 2
        while j < len(rows):
            rr = rows[j]
            suc = str(rr[0]).strip() if rr[0] else ''
            if not suc: j += 1; continue
            if ('2026' in suc and rr[1] is None) or 'Marino' in suc: break  # nuevo bloque
            eid = ENTITY.get(suc)
            if eid is None:
                print(f'-- AVISO: sucursal no mapeada: {suc!r}', file=sys.stderr); j += 1; continue
            for ci in colmes:
                if ci < len(rr):
                    for f, ffin in parse_cell(rr[ci]):
                        events.append((eid, tipo, f, ffin))
            j += 1
        i = j
    else:
        i += 1

print("-- seed lmq_agenda generado desde Cronograma_Mantenimientos_2026.xlsx")
print("DELETE FROM lmq_agenda WHERE descripcion = 'Importado del Cronograma 2026';")
for eid, tipo, f, ffin in events:
    ffin_sql = f"'{ffin}'" if ffin else 'NULL'
    tipo_sql = tipo.replace("'", "''")
    print(f"INSERT INTO lmq_agenda (entities_id, tipo, clase, fecha, fecha_fin, estado, descripcion)"
          f" VALUES ({eid}, '{tipo_sql}', 'preventivo', '{f}', {ffin_sql}, 'programado', 'Importado del Cronograma 2026');")
print(f'-- total eventos: {len(events)}', file=sys.stderr)
